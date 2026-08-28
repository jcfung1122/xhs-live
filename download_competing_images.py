#!/usr/bin/env python3
"""竞品商品图 + 合作买手表下载器

穿透商家（看同行）后:
1. 商品按【销售额降序 + 合作买手数降序】排序, 先展示列表供确认, 再下载图片
2. 首次下载时, 同时生成该店【合作买手表】(按本店合作销售额降序) 到店铺文件夹

用法:
    python download_competing_images.py <seller_id> <店铺名> [数量]
    python download_competing_images.py <seller_id> <店铺名> --list    # 仅展示排序列表, 不下载
    python download_competing_images.py <seller_id> <店铺名> --all     # 跳过确认, 全部下载
    python download_competing_images.py <seller_id> <店铺名> --no-buyers  # 不生成买手表

保存: <skill_dir>/竞品/<店铺名>/
  图片: <商品名>_¥价格_佣金XX%.{jpg|png}
  表格: <店铺名>_合作买手.xlsx
"""
import os
import sys
import re
import json
import requests

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
COMPETITOR_DIR = os.path.join(SKILL_DIR, "竞品")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Referer": "https://www.xiaohongshu.com/",
}


def sanitize_filename(name):
    name = re.sub(r'[\\/:*?"<>|\r\n\t]', "_", name)
    name = name.strip(" ._")
    return name[:60] if name else "unnamed"


def download_image(url, timeout=20):
    """下载图片, 返回 (bytes, ext) 或 (None, None)"""
    if url.startswith("//"):
        url = "https:" + url
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code == 200 and resp.content and len(resp.content) > 1000:
            ct = resp.headers.get("content-type", "").lower()
            ext = "png" if "png" in ct else ("webp" if "webp" in ct else "jpg")
            return resp.content, ext
    except Exception as e:
        print(f"    [WARN] 下载失败: {e}")
    return None, None


def fmt_sales(v):
    """格式化销售额: 万/亿"""
    try:
        v = float(v)
        if v >= 1e8:
            return f"{v/1e8:.1f}亿"
        if v >= 1e4:
            return f"{v/1e4:.1f}万"
        return f"{v:.0f}"
    except Exception:
        return str(v)


def sort_products(products):
    """按 销售额降序 + 合作买手数降序 排序"""
    def key(p):
        return (float(p.get("sales") or 0), int(p.get("coo_buyer_num") or 0))
    return sorted(products, key=key, reverse=True)


def show_product_list(products, title="商品列表"):
    """展示排序后的商品列表"""
    print("")
    print("=== " + title + " (按销售额+合作买手数降序) ===")
    print(f"{'序号':<4} {'商品名':<40} {'价格':>8} {'销售额':>10} {'买手数':>6} {'佣金':>6}")
    print("-" * 80)
    for i, p in enumerate(products, 1):
        name = (p.get("prodcut_name") or "")[:36]
        price = f"¥{float(p.get('price') or 0):.0f}"
        sales = fmt_sales(p.get("sales") or 0)
        buyers = p.get("coo_buyer_num") or 0
        comm = f"{float(p.get('commission') or 0)*100:.0f}%"
        print(f"{i:<4} {name:<40} {price:>8} {sales:>10} {buyers:>6} {comm:>6}")


def fetch_all_products(skill, seller_id, max_count):
    """分页拉取商品"""
    all_products = []
    page = 1
    while len(all_products) < max_count and page <= 10:
        res = skill.get_seller_products(seller_id, page=page, size=20)
        prods = res.get("products", [])
        if not prods:
            break
        all_products.extend(prods)
        total = res.get("total", 0)
        if len(all_products) >= total:
            break
        page += 1
    return all_products[:max_count]


def export_buyers_xlsx(skill, seller_id, shop_dir, shop_name):
    """生成合作买手表 (按本店合作销售额降序)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [WARN] 未安装 openpyxl, 跳过买手表生成 (pip install openpyxl)")
        return None

    all_buyers = []
    page = 1
    while page <= 30:
        res = skill.get_seller_coo_buyers(seller_id, page=page, size=20)
        buyers = res.get("buyers", [])
        if not buyers:
            break
        all_buyers.extend(buyers)
        total = res.get("total", 0)
        if len(all_buyers) >= total:
            break
        page += 1

    if not all_buyers:
        print("  [WARN] 未获取到合作买手")
        return None

    all_buyers.sort(key=lambda b: float(b.get("distributor_seller_data_info", {}).get("seller_sales_value") or 0), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合作买手"

    headers = ["排名", "达人昵称", "粉丝数", "城市", "场均销售", "本店合作销售额", "本店合作商品数", "是否新买手", "愿接受佣金"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, b in enumerate(all_buyers, 1):
        di = b.get("distributor_data_info", {})
        sd = b.get("distributor_seller_data_info", {})
        ex = b.get("distributor_extra_info", {})
        ws.append([
            i,
            di.get("distributor_name", ""),
            di.get("fans_num", 0),
            di.get("city", ""),
            di.get("avg_sale_amount", ""),
            float(sd.get("seller_sales_value") or 0),
            sd.get("coo_product_num", 0),
            "新买手" if sd.get("is_new_buyer") else "",
            ex.get("willing_commsion", ""),
        ])

    widths = [6, 24, 12, 14, 12, 18, 16, 10, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w
    ws.freeze_panes = "A2"

    xlsx_path = os.path.join(shop_dir, f"{sanitize_filename(shop_name)}_合作买手.xlsx")
    wb.save(xlsx_path)
    print(f"  [买手表] 已生成: {xlsx_path} ({len(all_buyers)} 位买手)")
    return xlsx_path


def main():
    args = sys.argv[1:]
    if len(args) < 2:
        print(__doc__)
        return 1

    seller_id = args[0]
    shop_name = args[1]
    only_list = "--list" in args
    skip_confirm = "--all" in args
    skip_buyers = "--no-buyers" in args
    count = 50
    for a in args[2:]:
        if a.isdigit():
            count = int(a)
            break

    sys.path.insert(0, SKILL_DIR)
    from pgy_live_skill import PgyLiveSkill, get_default_cookie_path

    skill = PgyLiveSkill(cookie_file=get_default_cookie_path())

    print(f"=== 穿透商家: {shop_name} (seller_id={seller_id}) ===")
    print(f"拉取商品 (上限 {count} 个)...")

    products = fetch_all_products(skill, seller_id, count)
    if not products:
        print("未获取到商品")
        return 1

    products = sort_products(products)
    show_product_list(products)

    if only_list:
        print(f"[--list] 仅展示, 未下载。共 {len(products)} 个商品。")
        return 0

    selected = products
    if not skip_confirm:
        print("")
        print("下载哪些? (回车=全部 / 输入序号如 1,3,5 / 输入范围 1-10 / n=取消)")
        choice = input("> ").strip().lower()
        if choice in ("n", "no", "q"):
            print("已取消")
            return 0
        if choice and choice != "all":
            idxs = set()
            for part in choice.split(","):
                part = part.strip()
                if "-" in part:
                    a, b = part.split("-")
                    idxs.update(range(int(a), int(b) + 1))
                elif part.isdigit():
                    idxs.add(int(part))
            selected = [products[i-1] for i in sorted(idxs) if 1 <= i <= len(products)]
            print(f"选择 {len(selected)} 个商品")

    shop_dir = os.path.join(COMPETITOR_DIR, sanitize_filename(shop_name))
    os.makedirs(shop_dir, exist_ok=True)

    if not skip_buyers:
        print("")
        print("[买手表] 生成该店合作买手表...")
        export_buyers_xlsx(skill, seller_id, shop_dir, shop_name)

    print(f"=== 下载 {len(selected)} 个商品图 ===")
    ok = 0
    fail = 0
    for i, p in enumerate(selected, 1):
        name = p.get("prodcut_name") or ""
        price = p.get("price") or 0
        commission = p.get("commission") or 0
        cover = p.get("product_cover") or ""

        if not cover or "/UNKNOWN" in cover or "unknown" in cover.lower():
            print(f"  [{i}/{len(selected)}] 平台无图: {name[:30]}")
            fail += 1
            continue

        price_str = f"¥{float(price):.0f}" if price else "无价"
        comm_str = f"佣金{float(commission)*100:.0f}%"
        data, ext = download_image(cover)
        if data is None:
            fail += 1
            print(f"  [{i}/{len(selected)}] FAIL: {name[:40]}")
            continue

        fname = sanitize_filename(f"{name}_{price_str}_{comm_str}.{ext}")
        fpath = os.path.join(shop_dir, fname)
        with open(fpath, "wb") as f:
            f.write(data)
        ok += 1
        print(f"  [{i}/{len(selected)}] OK: {fname[:65]}")

    print("")
    print(f"=== 完成: 图片成功 {ok}, 失败 {fail} ===")
    print(f"保存位置: {shop_dir}")
    return 0 if fail == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
